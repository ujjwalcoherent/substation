"""
Extract market data from Excel into JSON files for the Substation dashboard.
Produces: value.json, volume.json, segmentation_analysis.json
"""
import openpyxl
import json
import os
import re

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'Dataset-Global Online Substation Monitoring System Market.xlsx')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')

# Geography headers in the Excel (rows with no data = geography section headers)
# The order they appear in the Excel Value/Volume sheets
GEOGRAPHY_ORDER = [
    'Global', 'North America', 'Europe', 'Asia Pacific', 'Latin America', 'Middle East & Africa',
    'U.S.', 'Canada',
    'U.K.', 'Germany', 'France', 'Italy', 'Spain', 'Russia', 'Rest of Europe',
    'China', 'India', 'Japan', 'South Korea', 'ASEAN', 'Australia', 'Rest of Asia Pacific',
    'Brazil', 'Mexico', 'Argentina', 'Rest of Latin America',
    'GCC', 'South Africa', 'Rest of Middle East & Africa'
]

# Segment types that appear under each geography
SEGMENT_TYPES = [
    'By Offering', 'By Functionality', 'By Deployment Model',
    'By Communication and Integration Layer', 'By Substation Type',
    'By Voltage Range', 'By End User'
]

# Known parent segments (have children) - map parent -> list of children
# Built from Excel structure analysis
SEGMENT_HIERARCHY = {
    'By Offering': {
        'Hardware': [
            'Intelligent Electronic Devices (IEDs)',
            'Sensors and Transducers',
            'Data Acquisition Units and RTUs',
            'Transformer Monitoring Systems',
            'Partial Discharge Monitoring Systems',
            'Communication Gateways and Edge Devices',
            'Others (Power Quality Monitoring Devices, Metering and Measurement Devices, Time Synchronization Devices, etc.)'
        ],
        'Software': [
            'Monitoring and Visualization Platforms',
            'Analytics and Fault Diagnosis Software',
            'Asset Management and CMMS Integrations',
            'Others (Cybersecurity Software for Substations, etc.)'
        ],
        'Services': [
            'Managed Services',
            'Professional Services'
        ]
    },
    'By Functionality': {
        'Condition and Asset Health Monitoring': [],
        'Event, Alarm, and Disturbance Monitoring': [],
        'Diagnostics and Root Cause Analysis': [],
        'Predictive Maintenance and Asset Performance Optimization': [],
        'Power Quality and Grid Stability Monitoring': []
    },
    'By Deployment Model': {
        'On-Premises Deployment': [],
        'Cloud-Based Deployment': [],
        'Hybrid Deployment': []
    },
    'By Communication and Integration Layer': {
        'Wired Communication': [],
        'Wireless Communication': []
    },
    'By Substation Type': {
        'Transmission Substations': [],
        'Distribution Substations': []
    },
    'By Voltage Range': {
        '110 to 220 kV': [],
        '220 to 400 kV': [],
        '400 to 765 kV': [],
        'Above 765 kV': []
    },
    'By End User': {
        'Electric Utilities': [],
        'Independent Power Producers and Grid Asset Owners': [],
        'Industrial Substation': {
            'Oil and Gas': [],
            'Mining, Metals': [],
            'Manufacturing': []
        },
        'Infrastructure': {
            'Railways': [],
            'Metros': [],
            'Airports': []
        },
        'Others (Data Centers, Defense, etc.)': []
    }
}

GEOGRAPHY_HIERARCHY = {
    "Global": ["North America", "Europe", "Asia Pacific", "Latin America", "Middle East & Africa"],
    "North America": ["U.S.", "Canada"],
    "Europe": ["U.K.", "Germany", "France", "Italy", "Spain", "Russia", "Rest of Europe"],
    "Asia Pacific": ["China", "India", "Japan", "South Korea", "ASEAN", "Australia", "Rest of Asia Pacific"],
    "Latin America": ["Brazil", "Mexico", "Argentina", "Rest of Latin America"],
    "Middle East & Africa": ["GCC", "South Africa", "Rest of Middle East & Africa"]
}

YEARS = list(range(2021, 2034))  # 2021-2033


def clean_name(name):
    """Clean segment/geography names - trim trailing spaces."""
    if name is None:
        return None
    return name.strip()


def parse_sheet(ws):
    """
    Parse a Value or Volume sheet into a dict structure:
    { geography: { segment_type: { parent: { child: { year: value } } } } }
    """
    result = {}
    current_geo = None
    current_segment_type = None
    current_parent = None  # For hierarchical segments (Hardware, Software, etc.)
    current_grandparent = None  # For 3-level hierarchy (Industrial Substation > Oil and Gas)

    # First pass: find all geography header rows and their row numbers
    geo_rows = {}
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True)):
        name = clean_name(row[0])
        if name and name in GEOGRAPHY_ORDER:
            has_data = any(row[j] is not None for j in range(1, 14))
            if not has_data:
                geo_rows[i + 3] = name

    # Second pass: parse data
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True)):
        row_num = i + 3
        name = clean_name(row[0])
        if not name or name == 'Row Labels':
            continue

        has_data = any(row[j] is not None for j in range(1, 14))

        # Check if this is a geography header (no data)
        if row_num in geo_rows:
            current_geo = name
            current_segment_type = None
            current_parent = None
            current_grandparent = None
            if current_geo not in result:
                result[current_geo] = {}
            continue

        if current_geo is None:
            continue

        # Check if this is a segment type header (has data, matches known segment types)
        # Also handle "By Country" which we skip
        if name == 'By Country':
            current_segment_type = '__by_country__'
            current_parent = None
            current_grandparent = None
            continue

        if name in SEGMENT_TYPES:
            current_segment_type = name
            current_parent = None
            current_grandparent = None
            # Don't store segment type total - it would be double counting
            continue

        if current_segment_type is None or current_segment_type == '__by_country__':
            # Skip "By Country" entries - countries are separate geographies
            continue

        if not has_data:
            continue

        # Extract year values
        year_data = {}
        for j, year in enumerate(YEARS):
            val = row[j + 1]  # columns B through N (index 1-13)
            if val is not None:
                try:
                    year_data[str(year)] = round(float(val), 2)
                except (ValueError, TypeError):
                    pass

        if not year_data:
            continue

        # Determine where in the hierarchy this segment goes
        seg_type_hierarchy = SEGMENT_HIERARCHY.get(current_segment_type, {})

        # Is this a level-1 parent? (e.g., Hardware, Software, Services under By Offering)
        if name in seg_type_hierarchy:
            current_parent = name
            current_grandparent = None
            # Store parent data
            if current_segment_type not in result[current_geo]:
                result[current_geo][current_segment_type] = {}
            result[current_geo][current_segment_type][name] = year_data
            continue

        # Is this a level-2 child of the current parent?
        if current_parent and current_parent in seg_type_hierarchy:
            parent_children = seg_type_hierarchy[current_parent]
            if isinstance(parent_children, list):
                # Simple list of children
                if name in parent_children:
                    if current_segment_type not in result[current_geo]:
                        result[current_geo][current_segment_type] = {}
                    if current_parent not in result[current_geo][current_segment_type]:
                        result[current_geo][current_segment_type][current_parent] = year_data
                    if not isinstance(result[current_geo][current_segment_type][current_parent], dict) or \
                       any(k.isdigit() for k in result[current_geo][current_segment_type][current_parent]):
                        # Parent is stored as year_data, need to restructure
                        pass
                    # Store child under parent
                    key = f"{current_parent}__children"
                    if key not in result[current_geo][current_segment_type]:
                        result[current_geo][current_segment_type][key] = {}
                    result[current_geo][current_segment_type][key][name] = year_data
                    continue
            elif isinstance(parent_children, dict):
                # Dict means this parent has sub-parents (like Industrial Substation > Oil and Gas)
                if name in parent_children:
                    current_grandparent = name
                    key = f"{current_parent}__children"
                    if current_segment_type not in result[current_geo]:
                        result[current_geo][current_segment_type] = {}
                    if key not in result[current_geo][current_segment_type]:
                        result[current_geo][current_segment_type][key] = {}
                    result[current_geo][current_segment_type][key][name] = year_data
                    continue
                # Check if it's a grandchild
                if current_grandparent and current_grandparent in parent_children:
                    grandchildren = parent_children[current_grandparent]
                    if isinstance(grandchildren, list) and name in grandchildren:
                        key = f"{current_parent}__children"
                        gkey = f"{current_grandparent}__children"
                        if current_segment_type not in result[current_geo]:
                            result[current_geo][current_segment_type] = {}
                        if key not in result[current_geo][current_segment_type]:
                            result[current_geo][current_segment_type][key] = {}
                        if gkey not in result[current_geo][current_segment_type][key]:
                            result[current_geo][current_segment_type][key][gkey] = {}
                        result[current_geo][current_segment_type][key][gkey][name] = year_data
                        continue

        # Flat segment (no parent, direct child of segment type)
        # e.g., Condition and Asset Health Monitoring under By Functionality
        if current_parent is None or name not in (seg_type_hierarchy.get(current_parent, []) if isinstance(seg_type_hierarchy.get(current_parent), list) else []):
            # Could be a new level-1 item or a child we missed
            # Check if it's a known level-1 item
            if name in seg_type_hierarchy:
                current_parent = name
                current_grandparent = None
                if current_segment_type not in result[current_geo]:
                    result[current_geo][current_segment_type] = {}
                result[current_geo][current_segment_type][name] = year_data
            else:
                # It's a child of current_parent
                if current_parent:
                    key = f"{current_parent}__children"
                    if current_segment_type not in result[current_geo]:
                        result[current_geo][current_segment_type] = {}
                    if key not in result[current_geo][current_segment_type]:
                        result[current_geo][current_segment_type][key] = {}
                    result[current_geo][current_segment_type][key][name] = year_data

    return result


def restructure_to_nested(parsed_data):
    """
    Convert the flat parsed structure into the nested JSON format needed by the dashboard.

    Input: { geo: { seg_type: { parent: year_data, parent__children: { child: year_data } } } }
    Output: { geo: { seg_type: { parent: { child: { year: val } } } } }
    """
    result = {}
    for geo, seg_types in parsed_data.items():
        result[geo] = {}
        for seg_type, segments in seg_types.items():
            result[geo][seg_type] = {}

            # Separate parents from children markers
            parents = {}
            children_map = {}
            for key, val in segments.items():
                if key.endswith('__children'):
                    parent_name = key.replace('__children', '')
                    children_map[parent_name] = val
                else:
                    parents[key] = val

            for parent_name, parent_data in parents.items():
                if parent_name in children_map:
                    # This parent has children - create nested structure
                    result[geo][seg_type][parent_name] = {}
                    children = children_map[parent_name]
                    for child_key, child_val in children.items():
                        if child_key.endswith('__children'):
                            # Grandchildren
                            grandparent = child_key.replace('__children', '')
                            if grandparent not in result[geo][seg_type][parent_name]:
                                # Find grandparent data
                                if grandparent in children:
                                    result[geo][seg_type][parent_name][grandparent] = {}
                            if isinstance(child_val, dict):
                                for gc_name, gc_data in child_val.items():
                                    if grandparent in result[geo][seg_type][parent_name] and isinstance(result[geo][seg_type][parent_name][grandparent], dict):
                                        result[geo][seg_type][parent_name][grandparent][gc_name] = gc_data
                        else:
                            result[geo][seg_type][parent_name][child_key] = child_val
                else:
                    # Leaf segment (no children) - store year data directly
                    result[geo][seg_type][parent_name] = parent_data

    return result


def build_segmentation_analysis():
    """Build the segmentation_analysis.json with structure + geography hierarchy."""
    seg = {
        "_geography_hierarchy": GEOGRAPHY_HIERARCHY,
        "Global": {}
    }

    # By Offering
    seg["Global"]["By Offering"] = {
        "Hardware": {
            "Intelligent Electronic Devices (IEDs)": {},
            "Sensors and Transducers": {},
            "Data Acquisition Units and RTUs": {},
            "Transformer Monitoring Systems": {},
            "Partial Discharge Monitoring Systems": {},
            "Communication Gateways and Edge Devices": {},
            "Others (Power Quality Monitoring Devices, Metering and Measurement Devices, Time Synchronization Devices, etc.)": {}
        },
        "Software": {
            "Monitoring and Visualization Platforms": {},
            "Analytics and Fault Diagnosis Software": {},
            "Asset Management and CMMS Integrations": {},
            "Others (Cybersecurity Software for Substations, etc.)": {}
        },
        "Services": {
            "Managed Services": {},
            "Professional Services": {}
        }
    }

    # By Functionality
    seg["Global"]["By Functionality"] = {
        "Condition and Asset Health Monitoring": {},
        "Event, Alarm, and Disturbance Monitoring": {},
        "Diagnostics and Root Cause Analysis": {},
        "Predictive Maintenance and Asset Performance Optimization": {},
        "Power Quality and Grid Stability Monitoring": {}
    }

    # By Deployment Model
    seg["Global"]["By Deployment Model"] = {
        "On-Premises Deployment": {},
        "Cloud-Based Deployment": {},
        "Hybrid Deployment": {}
    }

    # By Communication and Integration Layer
    seg["Global"]["By Communication and Integration Layer"] = {
        "Wired Communication": {},
        "Wireless Communication": {}
    }

    # By Substation Type
    seg["Global"]["By Substation Type"] = {
        "Transmission Substations": {},
        "Distribution Substations": {}
    }

    # By Voltage Range
    seg["Global"]["By Voltage Range"] = {
        "110 to 220 kV": {},
        "220 to 400 kV": {},
        "400 to 765 kV": {},
        "Above 765 kV": {}
    }

    # By End User
    seg["Global"]["By End User"] = {
        "Electric Utilities": {},
        "Independent Power Producers and Grid Asset Owners": {},
        "Industrial Substation": {
            "Oil and Gas": {},
            "Mining, Metals": {},
            "Manufacturing": {}
        },
        "Infrastructure": {
            "Railways": {},
            "Metros": {},
            "Airports": {}
        },
        "Others (Data Centers, Defense, etc.)": {}
    }

    return seg


def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Parse Value sheet
    print("Parsing Value sheet...")
    value_parsed = parse_sheet(wb['Value'])
    value_data = restructure_to_nested(value_parsed)

    # Parse Volume sheet
    print("Parsing Volume sheet...")
    volume_parsed = parse_sheet(wb['Volume'])
    volume_data = restructure_to_nested(volume_parsed)

    # Build segmentation analysis
    print("Building segmentation analysis...")
    seg_data = build_segmentation_analysis()

    # Write output files
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    value_path = os.path.join(OUTPUT_DIR, 'value.json')
    print(f"Writing {value_path}")
    with open(value_path, 'w') as f:
        json.dump(value_data, f, indent=2)

    volume_path = os.path.join(OUTPUT_DIR, 'volume.json')
    print(f"Writing {volume_path}")
    with open(volume_path, 'w') as f:
        json.dump(volume_data, f, indent=2)

    seg_path = os.path.join(OUTPUT_DIR, 'segmentation_analysis.json')
    print(f"Writing {seg_path}")
    with open(seg_path, 'w') as f:
        json.dump(seg_data, f, indent=2)

    # Print summary
    print("\n=== Summary ===")
    print(f"Value: {len(value_data)} geographies")
    for geo in value_data:
        seg_types = list(value_data[geo].keys())
        print(f"  {geo}: {len(seg_types)} segment types: {seg_types}")
    print(f"\nVolume: {len(volume_data)} geographies")
    for geo in volume_data:
        seg_types = list(volume_data[geo].keys())
        print(f"  {geo}: {len(seg_types)} segment types")

    print("\nDone!")


if __name__ == '__main__':
    main()
